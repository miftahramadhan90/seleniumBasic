from tkinter import N
import openpyxl

'''Konsep Awal membaca cell di dalam excel '''
book = openpyxl.load_workbook('D:\\Belajar_Python\\2_Automation_Selenium\\a_UI_Basic_Selenium\\24_excel_handling\\data1.xlsx')
sheet1 = book['Sheet1']
sheet2 = book['Sheet2']

# cara 1 --> kelebihan cara 1 saat selesai ketik (1, 2). = auto suggestive function bawaan dari cell nya | utk awal sebaiknya pakai ini
print(sheet1.cell(2, 3).value)

# cara 2 :
print(sheet1['C2'].value)

#cek panjang baris dan pjg kolom yg terisi dari suatu Sheet :
print(sheet1.max_row) 
print(sheet1.max_column)

#mengubah data di cell (tanpa mengubah di excel)
# sheet1.cell(2,3).value = 'write 1'
# print(sheet1.cell(2,3).value)

#Contoh mengecek kondisional isi cell tertentu
if sheet1.cell(2,3).value == 'Ramadhan' :
    print('isi cell adalah : Ramadhan')
else:
    print('isi cell bukan Ramadhan tapi ' + sheet1.cell(2, 3).value)

#Teknik looping data1.xlsx
'''print SMUA DATA di 1 KOLOM tertentu'''
for i in range(2, sheet1.max_row+1):
    print(sheet1.cell(i, 1).value)

'''print DATA di 1 KOLOM tertentu SAMPAI di BARIS TERTENTU (Misal sampai baris TC2 aja) :'''
for i in range(2, 4):
    print(sheet1.cell(i, 1).value)
print('\n')
'''prinrt DATA di 1 baris dan semua kolom scr berurutan : baris-n1 kolom-n1, baris-n1 kolom-n2, baris-n1 kolom-n3, baris-n1 kolom-n++  '''
for i in range(2, sheet1.max_row+1):                # to get rows
    for j in range(2, sheet1.max_column+1):         # to get columns
        print(sheet1.cell(i, j).value)
print('\n')

'''print data hanya di TC tertentu menggunakan for loop dan kondisional if'''
for i in range(2, sheet1.max_row+1):
    if sheet1.cell(i, 1).value == "TC1- input1":
    
        for j in range(2, sheet1.max_column+1): #for di dlm for ini utk looping melebar ngabisin ke kolomnya dlu 
            print(sheet1.cell(i, j).value)
        break
    else:
        print("tidak ada nama testcaste tsb")
        break



